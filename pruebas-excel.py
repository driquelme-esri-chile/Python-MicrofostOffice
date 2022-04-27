
from asyncore import write
import datetime
import pandas as pd
import locale

from openpyxl import load_workbook
import pandas.io.formats.excel
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill,Border,Side



Analisis = ["lala","lele","lili","lolo","lulu","loli"]
Nombre = ["rala","lrje","dooli","jilo","lupuuu","lueli"]
Te = ["malm","meme","cicc","fod","ruu","boois"]
Est = ["kjajkla","oper","oisfi","oofo","uuu","ioli"]
Concesionario = ["oieg","yatsd","idtg","fgkj","ysadu","mami"]
Area = ["lala","lele","lili","lolo","lulu","loli"]
Rol = ["lala","lele","lili","lolo","lulu","loli"]

    
df = pd.DataFrame({
    'Analisis': Analisis,
    'Nombre': Nombre,
    'T': Te,
    'Est': Est,
    'Concesionario': Concesionario,
    'Area': Area,
    'Rol': Rol
})



book = load_workbook('templates/templates_ads.xlsx')

outputFileExcel: str = jobsDir + "\\" + 'Analisis de Sobreposicion {}.xlsx'.format(nombre_salida)

writer = pd.ExcelWriter(outputFileExcel,  engine='openpyxl') 

print(writer)

writer.book = book

print(book)

pandas.io.formats.excel.header_style = None



writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df.to_excel(writer, book.worksheets[0].title, startcol = 0,  index = False)


sheet = writer.sheets['Analisis']

lista = ['A','B','C','D','E','G','H','I','M','N','O']

for x in range(0,len(lista)):
    for cell in sheet[lista[x]]:
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),  top=Side(style='thin'), bottom=Side(style='thin')) 


for cell in sheet["L"]:
    cell.alignment = Alignment(horizontal="right")
    cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),  top=Side(style='thin'), bottom=Side(style='thin')) 
for cell in sheet["F"]:
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '#,##0.00'
    cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),  top=Side(style='thin'), bottom=Side(style='thin')) 
for cell in sheet["J"]:
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '#,##0.00'
    cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),  top=Side(style='thin'), bottom=Side(style='thin')) 
for cell in sheet["K"]:
    cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),  top=Side(style='thin'), bottom=Side(style='thin')) 



writer.save()
arcpy.SetParameterAsText(1, writer)